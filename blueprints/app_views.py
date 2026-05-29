"""
RegHub 360 — 앱 페이지 Blueprint (MVP 6개 화면)
  /app/search              통합 검색
  /app/product/<code>      의약품 Product 360
  /app/product-food/<code> 식품 Product 360 (자리 표시자 + NO 1·153 부분)
  /app/timeline            통합 이벤트 타임라인 (회수+처분+안전성서한+공급)
  /app/copilot             AI 코파일럿

  ※ /app/(overview), /app/workspace, /app/watchlist, /app/reports
    Phase 2로 미루어 라우트 제거 (사이드바에서는 비활성 표시).
"""
import io
import logging
from flask import Blueprint, render_template, request, redirect, url_for, send_file, jsonify

from ..api_client import (
    fetch_approval, fetch_approval_detail,
    fetch_disciplinary, fetch_recall, fetch_identification,
)
from ..api_extras import (
    # 핵심 11
    fetch_drug_easy, fetch_drug_bundle, fetch_drug_safety_letter,
    fetch_drug_supply_stop, fetch_drug_supply_lack,
    fetch_drug_gmp, fetch_dur_item, fetch_dur_ingredient,
    fetch_food_nutrition,
    # 확장 R&D 8개
    fetch_drug_patent, fetch_drug_lawsuit, fetch_drug_fda_p4, fetch_drug_fda_orangebook,
    fetch_drug_clinical, fetch_drug_clinical_org, fetch_drug_reference, fetch_drug_bioeq,
    # 확장 품질·인증 3개
    fetch_drug_release, fetch_drug_entity_list, fetch_drug_dmf,
    # 확장 재심사·희귀 3개
    fetch_drug_review, fetch_drug_reeval, fetch_drug_orphan,
    # 확장 식품·건기식 2개
    fetch_food_inspect, fetch_hf_gmp,
)


def _client_filter(api_fn, fields, q, n=50, **kwargs):
    """fetch 함수 호출 + 클라이언트 사이드 필터링.
    검색 파라미터 미지원 API용. needle 이 fields 중 하나라도 포함되면 통과.
    """
    if not q:
        return []
    try:
        kwargs.setdefault("num_of_rows", n)
        r = api_fn(**kwargs)
        needle = q.lower()
        return [it for it in (r.get("items") or [])
                if any(needle in str(it.get(f, "")).lower() for f in fields)]
    except Exception as e:
        logger.warning(f"_client_filter {api_fn.__name__} 실패: {e}")
        return []


def _parallel_filter(jobs, max_workers=8):
    """여러 _client_filter 호출을 병렬 실행 — 응답 시간 단축.
    jobs: list of (label, api_fn, fields, q, n) tuples
    return: dict[label] = filtered_items
    """
    from concurrent.futures import ThreadPoolExecutor
    results = {}
    def _run(label, fn, fields, q, n):
        return label, _client_filter(fn, fields, q, n=n)
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = [ex.submit(_run, *j) for j in jobs]
        for fut in futs:
            try:
                label, items = fut.result(timeout=20)
                results[label] = items
            except Exception as e:
                logger.warning(f"_parallel_filter 실패: {e}")
    return results


def _filter_items(items, q, fields):
    """클라이언트 사이드 필터링 — 응답에 검색 파라미터가 없을 때 사용.
    fields 리스트 중 하나라도 query 부분일치하면 통과."""
    if not q:
        return items
    needle = q.strip().lower()
    out = []
    for it in items:
        for f in fields:
            v = it.get(f)
            if v and needle in str(v).lower():
                out.append(it)
                break
    return out

logger = logging.getLogger(__name__)

app_bp = Blueprint("app", __name__, url_prefix="/app")


# ────────────────────────────────────────────────────────────────────────────
# Search (호환 리다이렉트 — / 가 메인이므로)
# ────────────────────────────────────────────────────────────────────────────

@app_bp.route("/")
def index():
    """앱 진입 — / 로 리다이렉트 (메인 검색 페이지)."""
    return redirect("/", code=302)


def _is_company(q: str) -> bool:
    """업체명 여부 휴리스틱."""
    if not q:
        return False
    return any(s in q for s in ("제약", "약품", "(주)", "주식회사", "Pharma", "광동", "동아", "한미", "유한", "제일"))


def _is_drug_name(q: str) -> bool:
    """의약품 제품명 여부 휴리스틱."""
    if not q:
        return False
    return any(suffix in q for suffix in ("정", "캡슐", "시럽", "주", "액", "연질", "환", "산제", "패치"))


@app_bp.route("/search")
def search():
    """
    통합 검색 화면 — q 파라미터가 있으면 실 API 호출 결과를 컨텍스트에 주입.
    템플릿은 search_results (없으면 디자인 mock 데이터 그대로 표시).
    """
    q = (request.args.get("q") or "").strip()
    ctx = {"query": q, "search_results": None, "active_page": "search"}

    if q:
        is_company = _is_company(q)
        is_drug = _is_drug_name(q)
        kind = "업체" if is_company else ("의약품" if is_drug else "제품")

        # 핵심 API 다중 호출 — NO 140 허가정보 + e약은요 + 회수 + 행정처분 + 안전성 + 공급 + 식품
        results = {"kind": kind}
        try:
            if is_company:
                results["approval"] = fetch_approval(entp_name=q, num_of_rows=10)
                results["drug_easy"] = fetch_drug_easy(entp_name=q, num_of_rows=6)
            else:
                results["approval"] = fetch_approval(item_name=q, num_of_rows=10)
                results["drug_easy"] = fetch_drug_easy(item_name=q, num_of_rows=6)
            # 회수·처분·안전성·공급·식품 = 검색 파라미터 미지원 → 충분히 큰 numOfRows로 받고 필터
            results["recall_raw"] = fetch_recall(num_of_rows=200)
            results["disc_raw"] = fetch_disciplinary(num_of_rows=200)
            results["safety"] = fetch_drug_safety_letter(num_of_rows=200)
            results["supply"] = fetch_drug_supply_stop(num_of_rows=200)
            results["food"] = fetch_food_nutrition(num_of_rows=200)
        except Exception as e:
            logger.error(f"search 데이터 수집 실패: {e}")

        # 회수·행정처분·안전성·공급·식품 모두 클라이언트 사이드 필터
        recalls_filtered = _filter_items(
            results.get("recall_raw", {}).get("items", []),
            q,
            fields=["PRDUCT", "ENTRPS", "ITEM_NAME", "ENTP_NAME"],
        )
        sanctions_filtered = _filter_items(
            results.get("disc_raw", {}).get("items", []),
            q,
            fields=["ENTP_NAME", "ITEM_NAME", "ENTRPS_NAME", "VIOLATION_DETAIL"],
        )
        safety_filtered = _filter_items(
            results.get("safety", {}).get("items", []),
            q,
            fields=["TITLE", "SUMMARY", "DETAIL_CN", "TBL_CN"],
        )
        supply_filtered = _filter_items(
            results.get("supply", {}).get("items", []),
            q,
            fields=["ENTP_NAME", "ITEM_NAME", "PRDUCT", "ENTRPS_NAME"],
        )
        food_filtered = _filter_items(
            results.get("food", {}).get("items", []),
            q,
            fields=["FOOD_NM_KR", "BSSH_NM", "MANUFACTURE"],
        )

        # 허가정보 = 가장 신뢰성 높은 제품 매칭 (NO 140)
        approval_items = results.get("approval", {}).get("items", [])

        # KPI 카운트 — 모두 정밀 필터링된 카운트
        counts = {
            "drug": results.get("approval", {}).get("totalCount", 0),
            "recall": len(recalls_filtered),
            "sanction": len(sanctions_filtered),
            "safety": len(safety_filtered),
            "supply": len(supply_filtered),
            "food": len(food_filtered),
        }

        # 표시용 products = 허가정보(우선) + e약은요(보조)
        products = approval_items[:6] if approval_items else \
                   results.get("drug_easy", {}).get("items", [])[:6]

        # top_products — 의약품 허가 결과 + Risk Score 정렬 (회수/처분 매칭 품목 우선)
        recalls_by_product = {}
        for r in recalls_filtered:
            key = (r.get("PRDUCT") or "").lower()
            if key:
                recalls_by_product[key] = recalls_by_product.get(key, 0) + 1
        sanctions_by_entity = {}
        for s in sanctions_filtered:
            key = (s.get("ENTP_NAME") or "").lower()
            if key:
                sanctions_by_entity[key] = sanctions_by_entity.get(key, 0) + 1

        scored_products = []
        for p in approval_items[:15]:
            name = (p.get("ITEM_NAME") or "").lower()
            entp = (p.get("ENTP_NAME") or "").lower()
            score = 0
            for k, v in recalls_by_product.items():
                if k and (k in name or name in k):
                    score += v * 40
            for k, v in sanctions_by_entity.items():
                if k and (k in entp or entp in k):
                    score += v * 30
            scored_products.append((score, p))
        scored_products.sort(key=lambda t: t[0], reverse=True)
        top_products = [
            {**p, "_risk": s, "_grade": "High" if s >= 60 else "Medium" if s >= 30 else "Low" if s > 0 else "Normal"}
            for s, p in scored_products[:6]
        ]

        # watch_events — 회수 + 처분 + 안전성서한 + 공급중단 통합 평면화 (시간순)
        watch_events = []
        for r in recalls_filtered[:8]:
            watch_events.append({
                "type": "회수",
                "severity_level": "HIGH", "severity_color": "warn",
                "title": (r.get("PRDUCT") or "(제품 미상)")[:50],
                "summary": (r.get("RTRVL_RESN") or "")[:60] + " · " + (r.get("ENTRPS") or ""),
                "date": (r.get("RECALL_COMMAND_DATE") or "")[:10],
            })
        for s in sanctions_filtered[:5]:
            watch_events.append({
                "type": "행정처분",
                "severity_level": "HIGH", "severity_color": "warn",
                "title": (s.get("ENTP_NAME") or "(업체 미상)")[:50],
                "summary": (s.get("VIOLATION_DETAIL") or s.get("VIOLATION_LAW") or "")[:60],
                "date": (s.get("DSPS_DCSNDT") or "")[:10],
            })
        for sl in (results.get("safety", {}).get("items", []) or [])[:3]:
            # 안전성서한 필터링 — 검색어 포함만
            title = sl.get("TITLE") or ""
            if q.lower() in title.lower() or q.lower() in (sl.get("SUMRY_CONT") or "").lower():
                watch_events.append({
                    "type": "안전성서한",
                    "severity_level": "HIGH", "severity_color": "warn",
                    "title": title[:50],
                    "summary": (sl.get("SUMRY_CONT") or "")[:60],
                    "date": (sl.get("PBANC_YMD") or "")[:10],
                })
        watch_events.sort(key=lambda e: e["date"], reverse=True)

        # ────────────────────────────────────────────────────────────────────
        # 확장 17개 API — 클라이언트 사이드 필터링 (검색어 매칭만)
        # ────────────────────────────────────────────────────────────────────
        rnd_items = {"patent": [], "lawsuit": [], "fda_p4": [], "orangebook": [],
                     "clinical": [], "reference": [], "bioeq": []}
        quality_items = {"gmp": [], "shipment": [], "entity": [], "dmf": []}
        review_items = {"review": [], "reeval": [], "orphan": []}
        inspection_items = []
        hf_gmp_items = []

        if q:
            # 모든 확장 API 병렬 호출 — 응답 시간 30초+ → 5~8초로 단축
            jobs = [
                ("rnd_patent",     fetch_drug_patent,        ["ITEM_NAME", "INGR_NAME", "ENTP_NAME"], q, 50),
                ("rnd_lawsuit",    fetch_drug_lawsuit,       ["PRT_NAME", "INGR_NAME"],             q, 30),
                ("rnd_fda_p4",     fetch_drug_fda_p4,        ["DRUG_NAME", "RLD"],                  q, 30),
                ("rnd_orangebook", fetch_drug_fda_orangebook,["PRT_NAME", "INGR_NAME"],             q, 30),
                ("rnd_clinical",   fetch_drug_clinical,      ["GOODS_NAME", "APPLY_ENTP_NAME", "CLINIC_EXAM_TITLE"], q, 50),
                ("rnd_reference",  fetch_drug_reference,     ["ITEM_NAME", "INGR_NAME", "ENTP_NAME"], q, 30),
                ("rnd_bioeq",      fetch_drug_bioeq,         ["ITEM_NAME", "INGR_KOR_NAME", "ENTP_NAME"], q, 30),
                ("q_shipment",     fetch_drug_release,       ["MANUF_ENTP_NAME", "GOODS_NAME"],     q, 30),
                ("q_entity",       fetch_drug_entity_list,   ["ENTRPS", "RPRSNTV"],                 q, 30),
                ("q_dmf",          fetch_drug_dmf,           ["ENTP_NAME", "MNFCTR_NAME", "INGR_KOR_NAME"], q, 30),
                ("rv_review",      fetch_drug_review,        ["ENTP_NAME", "ITEM_NAME"],            q, 30),
                ("rv_reeval",      fetch_drug_reeval,        ["ENTP_NAME", "ITEM_NAME"],            q, 30),
                ("rv_orphan",      fetch_drug_orphan,        ["PRDT_NAME", "PRODT_NAME", "TARGET_DISEASE"], q, 30),
                ("inspection",     fetch_food_inspect,       ["PRDUCT", "ENTRPS", "IMPROPT_ITM"],   q, 30),
                ("hf_gmp",         fetch_hf_gmp,             ["BSSH_NM", "PRSDNT_NM"],              q, 20),
            ]
            par = _parallel_filter(jobs, max_workers=10)
            # 결과 분배
            rnd_items["patent"]     = par.get("rnd_patent", [])
            rnd_items["lawsuit"]    = par.get("rnd_lawsuit", [])
            rnd_items["fda_p4"]     = par.get("rnd_fda_p4", [])
            rnd_items["orangebook"] = par.get("rnd_orangebook", [])
            rnd_items["clinical"]   = par.get("rnd_clinical", [])
            rnd_items["reference"]  = par.get("rnd_reference", [])
            rnd_items["bioeq"]      = par.get("rnd_bioeq", [])
            quality_items["shipment"] = par.get("q_shipment", [])
            quality_items["entity"]   = par.get("q_entity", [])
            quality_items["dmf"]      = par.get("q_dmf", [])
            review_items["review"] = par.get("rv_review", [])
            review_items["reeval"] = par.get("rv_reeval", [])
            review_items["orphan"] = par.get("rv_orphan", [])
            inspection_items = par.get("inspection", [])
            hf_gmp_items = par.get("hf_gmp", [])

        # 합계 카운트
        rnd_total = sum(len(v) for v in rnd_items.values())
        quality_total = sum(len(v) for v in quality_items.values())
        review_total = sum(len(v) for v in review_items.values())

        # 확장 카운트 통합
        counts.update({
            "rnd_total": rnd_total,
            "rnd": {k: len(v) for k, v in rnd_items.items()},
            "quality_total": quality_total,
            "quality": {k: len(v) for k, v in quality_items.items()},
            "review_total": review_total,
            "review": {k: len(v) for k, v in review_items.items()},
            "inspection": len(inspection_items),
            "hf_gmp": len(hf_gmp_items),
        })

        # 실제로 응답을 받은 API 카테고리 수 (성공/0건 모두 카운트, 실패는 제외)
        core_called = sum(1 for k in ("approval", "drug_easy", "recall_raw",
                                       "disc_raw", "safety", "supply", "food")
                          if results.get(k))
        ext_called = len(par) if q else 0   # 확장 API job 결과 dict (실패는 _parallel_filter 에서 제외됨)
        api_count_dyn = core_called + ext_called

        ctx["search_results"] = {
            "kind": kind,
            "counts": counts,
            "api_count": api_count_dyn,  # 실 호출 성공 API 수
            "products": products,
            "top_products": top_products,
            "watch_events": watch_events[:6],
            "recalls": recalls_filtered[:3],
            "sanctions": sanctions_filtered[:3],
            # 확장 신규 카드 데이터
            "rnd_items": rnd_items,
            "quality_items": quality_items,
            "review_items": review_items,
            "inspection_items": inspection_items[:3],
            "hf_gmp_items": hf_gmp_items[:3],
        }

    return render_template("app/search.html", **ctx)


# ────────────────────────────────────────────────────────────────────────────
# Product 360 (의약품 / 식품 / 향후 의약외품·건강기능식품)
# ────────────────────────────────────────────────────────────────────────────

@app_bp.route("/product/<code>")
def product_drug(code):
    """
    의약품 Product 360 — 8개 탭.
    NO 140(허가) + 248(e약은요) + 563(낱알식별) + 539(회수) + 547(안전성서한)
    + 269(묶음의약품/HIRA·ATC) + 531(DUR 품목) + 534(공급중단) 실 데이터 주입.
    """
    product = None
    drug_easy = None
    identification = None
    related_recalls = []
    same_entity_recalls = []  # 동일 업체 다른 품목 회수 (참고용)
    safety_letters = []
    bundle = None
    dur_items = []
    supply_stops = []
    ingredient_kr = None  # 한글 주성분명
    detail = None  # NO 140 상세 응답 (효능효과·용법용량·재심사일·ATC 등)
    # 확장 R&D / 품질 / 재심사 데이터
    p_patents = []
    p_lawsuits = []
    p_fda_p4 = []
    p_orangebook = []
    p_clinical = []
    p_reference = []
    p_bioeq = []
    p_release = []
    p_dmf = []
    p_review = []
    p_reeval = []
    p_orphan = []

    try:
        # 1. NO 140 의약품 허가정보 — code 는 edi_code(보험코드) / item_name 순서로 시도
        # ※ item_seq 파라미터는 NO 140에서 작동하지 않음 (전수 검증 완료)
        ap = fetch_approval(edi_code=code, num_of_rows=1)
        if ap.get("totalCount", 0) == 0 or len(ap.get("items", [])) == 0:
            # edi_code 매칭 실패 → item_name 매칭 시도
            ap = fetch_approval(item_name=code, num_of_rows=1)
        items = ap.get("items", [])
        if items:
            product = items[0]
            item_name = product.get("ITEM_NAME") or ""
            entp_name = product.get("ENTP_NAME") or ""
            ingr_name = product.get("ITEM_INGR_NAME") or ""

            # 2. NO 248 e약은요 — itemSeq 카멜케이스 (작동 확인)
            try:
                ed = fetch_drug_easy(item_seq=code, num_of_rows=1)
                if ed.get("items"):
                    drug_easy = ed["items"][0]
            except Exception:
                pass

            # 2-1. NO 140 상세정보 (getDrugPrdtPrmsnDtlInq06)
            # — 효능효과·용법용량·재심사일·ATC·CHART·VALID_TERM·BAR_CODE 등 42 필드
            # 다중 검색 전략으로 일시 실패 대응: item_seq → edi_code → item_name → entp_name
            detail_attempts = []
            item_seq = (product or {}).get("ITEM_SEQ") or ""
            detail_queries = [
                ("item_seq", {"item_seq": item_seq}) if item_seq else None,
                ("edi_code", {"edi_code": code}) if code else None,
                ("item_name", {"item_name": item_name}) if item_name else None,
                ("item_name(prefix)", {"item_name": item_name.split('(')[0]}) if item_name and '(' in item_name else None,
            ]
            for label, q in [x for x in detail_queries if x]:
                try:
                    dr = fetch_approval_detail(**q, num_of_rows=1)
                    detail_attempts.append(
                        f"{label}={list(q.values())[0]!r} → success={dr.get('success')}, items={len(dr.get('items', []))}"
                    )
                    if dr.get("items"):
                        detail = dr["items"][0]
                        logger.info(f"NO 140 상세 성공 [{label}]: {len(detail)} fields")
                        break
                except Exception as e:
                    detail_attempts.append(f"{label} → EXCEPTION: {e}")
            if not detail:
                logger.warning(
                    "NO 140 상세 조회 실패 — 시도: " + " | ".join(detail_attempts)
                )

            # 3. NO 563 낱알식별 — item_name (검증된 파라미터)
            try:
                idr = fetch_identification(item_name=item_name.split('(')[0], num_of_rows=1)
                if idr.get("items"):
                    identification = idr["items"][0]
            except Exception:
                pass

            # 3-1. 한글 주성분명 추출 — ITEM_NAME 안 괄호 내용 우선 사용
            #   "베니톨정(미세정제플라보노이드분획물)" → "미세정제플라보노이드분획물"
            import re as _re
            m_ingr = _re.search(r'\(([^()]+)\)', item_name or '')
            if m_ingr:
                ingredient_kr = m_ingr.group(1).strip()

            # 4. NO 539 회수 — 파라미터 무시되므로 클라이언트 필터
            # ⚠️ 핵심 규칙: 이 "제품"의 회수이력 (제품명 매칭) 만 카운트.
            #    동일 업체 다른 품목 회수는 사용자가 베니톨정 검색 시 보고 싶지 않음 (v3 §1.2).
            try:
                rr = fetch_recall(num_of_rows=200)
                needle_n = item_name.split('(')[0].strip().lower() if item_name else ""
                if needle_n and len(needle_n) >= 2:
                    related_recalls = [
                        it for it in rr.get("items", [])
                        if needle_n in str(it.get("PRDUCT", "")).lower()
                    ][:5]
                # 별도: 동일 업체 다른 품목 회수 (참고 표시용, related_recalls와 분리)
                needle_e = (entp_name or "").strip().lower()
                if needle_e and len(needle_e) >= 2:
                    same_entity_recalls = [
                        it for it in rr.get("items", [])
                        if needle_e in str(it.get("ENTRPS", "")).lower()
                        and (not needle_n or needle_n not in str(it.get("PRDUCT", "")).lower())
                    ][:5]
                else:
                    same_entity_recalls = []
            except Exception:
                same_entity_recalls = []

            # 5. NO 547 안전성서한 — TITLE (대문자) 검증 완료
            try:
                sl = fetch_drug_safety_letter(title=item_name.split('(')[0], num_of_rows=3)
                safety_letters = sl.get("items", [])
            except Exception:
                pass

            # 6. NO 269 묶음의약품 — trustItemName 검증 완료
            try:
                bd = fetch_drug_bundle(item_name=item_name.split('(')[0], num_of_rows=1)
                if bd.get("items"):
                    bundle = bd["items"][0]
            except Exception:
                pass

            # 7. NO 531 DUR 품목 (병용금기) — itemName 카멜케이스 검증
            try:
                from ..api_extras import fetch_dur_item
                du = fetch_dur_item(category="combine", item_name=item_name.split('(')[0], num_of_rows=3)
                dur_items = du.get("items", [])
            except Exception:
                pass

            # 8. NO 534 공급중단 — 제품명 매칭만 (회수와 동일 정책)
            try:
                ss = fetch_drug_supply_stop(num_of_rows=200)
                needle_n2 = item_name.split('(')[0].strip().lower() if item_name else ""
                if needle_n2 and len(needle_n2) >= 2:
                    supply_stops = [
                        it for it in ss.get("items", [])
                        if needle_n2 in str(it.get("ITEM_NAME", "")).lower()
                    ][:3]
            except Exception:
                pass

            # ────────────────────────────────────────────────────────────
            # 확장 R&D / 품질 / 재심사 — 제품명/성분명 기반 클라이언트 필터링
            # ────────────────────────────────────────────────────────────
            name_short = item_name.split('(')[0]
            try:
                p_patents = _client_filter(
                    fetch_drug_patent, ["ITEM_NAME", "INGR_NAME", "ENTP_NAME"], name_short, n=50
                )[:5]
            except Exception: pass
            try:
                p_lawsuits = _client_filter(
                    fetch_drug_lawsuit, ["PRT_NAME", "INGR_NAME"], name_short, n=30
                )[:3]
            except Exception: pass
            try:
                if ingredient_kr or product.get("ITEM_INGR_NAME"):
                    ingr_q = ingredient_kr or product.get("ITEM_INGR_NAME", "")
                    p_fda_p4 = _client_filter(
                        fetch_drug_fda_p4, ["DRUG_NAME"], ingr_q.split()[0] if ingr_q else "", n=30
                    )[:3]
                    p_orangebook = _client_filter(
                        fetch_drug_fda_orangebook, ["INGR_NAME", "PRT_NAME"], ingr_q.split()[0] if ingr_q else "", n=30
                    )[:3]
            except Exception: pass
            try:
                p_clinical = _client_filter(
                    fetch_drug_clinical, ["GOODS_NAME", "APPLY_ENTP_NAME"], name_short, n=50
                )[:5]
            except Exception: pass
            try:
                p_reference = _client_filter(
                    fetch_drug_reference, ["ITEM_NAME", "INGR_NAME", "ENTP_NAME"], name_short, n=30
                )[:3]
            except Exception: pass
            try:
                p_bioeq = _client_filter(
                    fetch_drug_bioeq, ["ITEM_NAME", "INGR_KOR_NAME", "ENTP_NAME"], name_short, n=30
                )[:3]
            except Exception: pass
            try:
                p_release = _client_filter(
                    fetch_drug_release, ["GOODS_NAME", "MANUF_ENTP_NAME"], name_short, n=30
                )[:3]
            except Exception: pass
            try:
                p_dmf = _client_filter(
                    fetch_drug_dmf, ["INGR_KOR_NAME", "ENTP_NAME", "MNFCTR_NAME"],
                    ingredient_kr or entp_name, n=30
                )[:3]
            except Exception: pass
            try:
                p_review = _client_filter(
                    fetch_drug_review, ["ITEM_NAME", "ENTP_NAME"], name_short, n=30
                )[:3]
            except Exception: pass
            try:
                p_reeval = _client_filter(
                    fetch_drug_reeval, ["ITEM_NAME", "ENTP_NAME"], name_short, n=30
                )[:3]
            except Exception: pass
            try:
                p_orphan = _client_filter(
                    fetch_drug_orphan, ["PRDT_NAME", "PRODT_NAME"], name_short, n=30
                )[:3]
            except Exception: pass
    except Exception as e:
        logger.warning(f"Product 360 데이터 수집 실패 (code={code}): {e}")

    # 자사 마스터 대조 (v2 §2 S1) — 자사 품목이면 배지 + 관련 QA 이벤트
    own_product = None
    own_events = []
    try:
        from .qa import lookup as qa_lookup
        seq = (product or {}).get("ITEM_SEQ") if product else None
        own_product = qa_lookup.match_own_product(
            item_seq=seq, edi_code=code, item_name=(product or {}).get("ITEM_NAME") if product else None)
        if own_product:
            own_events = qa_lookup.events_for_item(own_product["item_code"])
    except Exception as e:
        logger.debug(f"자사 마스터 대조 실패: {e}")

    return render_template(
        "app/product_drug.html",
        active_page="search",
        product_code=code,
        product_live=product,
        drug_easy=drug_easy,
        identification=identification,
        detail=detail,
        own_product=own_product,
        own_events=own_events,
        related_recalls=related_recalls,
        same_entity_recalls=same_entity_recalls,
        safety_letters=safety_letters,
        bundle=bundle,
        dur_items=dur_items,
        supply_stops=supply_stops,
        ingredient_kr=ingredient_kr,
        # 확장 R&D / 품질 / 재심사
        p_patents=p_patents,
        p_lawsuits=p_lawsuits,
        p_fda_p4=p_fda_p4,
        p_orangebook=p_orangebook,
        p_clinical=p_clinical,
        p_reference=p_reference,
        p_bioeq=p_bioeq,
        p_release=p_release,
        p_dmf=p_dmf,
        p_review=p_review,
        p_reeval=p_reeval,
        p_orphan=p_orphan,
    )


@app_bp.route("/product-quasi/<code>")
def product_quasi(code):
    """의약외품 Product 360 — NO 145 + 564(처분) + 547(안전성서한) + 534(공급) 통합."""
    from ..api_extras import (fetch_quasi_approval, fetch_drug_safety_letter,
                              fetch_drug_supply_stop)
    product = None
    safety_letters = []
    related_recalls = []
    related_sanctions = []
    supply_stops = []
    try:
        # NO 145 의약외품 허가 — 검색 키워드가 품목명 우선
        resp = fetch_quasi_approval(item_name=code, num_of_rows=1)
        if resp.get("items"):
            product = resp["items"][0]
            item_name = product.get("ITEM_NAME") or ""
            entp_name = product.get("ENTP_NAME") or ""
            # 안전성서한
            try:
                sl = fetch_drug_safety_letter(title=item_name.split('(')[0], num_of_rows=3)
                safety_letters = sl.get("items", [])
            except Exception:
                pass
            # 행정처분 (클라이언트 필터)
            try:
                dr = fetch_disciplinary(num_of_rows=200)
                needle = (entp_name or item_name).lower()
                related_sanctions = [
                    it for it in dr.get("items", [])
                    if needle and (needle in str(it.get("ENTP_NAME", "")).lower()
                                   or needle in str(it.get("ITEM_NAME", "")).lower())
                ][:5]
            except Exception:
                pass
            # 공급중단 (클라이언트 필터)
            try:
                ss = fetch_drug_supply_stop(num_of_rows=200)
                needle = (item_name.split('(')[0] or "").lower()
                supply_stops = [
                    it for it in ss.get("items", [])
                    if needle and needle in str(it.get("ITEM_NAME", "")).lower()
                ][:3]
            except Exception:
                pass
    except Exception as e:
        logger.warning(f"product_quasi 데이터 수집 실패 (code={code}): {e}")

    return render_template(
        "app/product_quasi.html",
        active_page="search",
        product_code=code,
        product=product,
        safety_letters=safety_letters,
        related_sanctions=related_sanctions,
        related_recalls=related_recalls,
        supply_stops=supply_stops,
    )


@app_bp.route("/product-food/<code>")
def product_food(code):
    """식품 Product 360 — 자리 표시자 (식품 API 5개 명세 대기)."""
    food_nutrition = None
    food_recall_import = None
    try:
        from ..api_extras import fetch_food_nutrition, fetch_food_recall_import
        # NO 1 영양 DB 검색
        nutri_r = fetch_food_nutrition(food_name=code, num_of_rows=1)
        if nutri_r.get("items"):
            food_nutrition = nutri_r["items"][0]
        # NO 153 수입식품 회수
        recall_r = fetch_food_recall_import(product_name=code, num_of_rows=1)
        if recall_r.get("items"):
            food_recall_import = recall_r["items"][0]
    except Exception as e:
        logger.warning(f"Product-food 데이터 수집 실패: {e}")

    return render_template(
        "app/product_food.html",
        active_page="search",
        product_code=code,
        food_nutrition=food_nutrition,
        food_recall_import=food_recall_import,
    )


# ────────────────────────────────────────────────────────────────────────────
# 통합 기능 페이지 (MVP 유지: timeline · copilot)
# ────────────────────────────────────────────────────────────────────────────

@app_bp.route("/timeline")
def timeline():
    """기존 /app/timeline → /app/monitor 301 redirect (KD-IRIS v3)."""
    return redirect(url_for("app.monitor"), code=301)


@app_bp.route("/monitor")
def monitor():
    """이벤트 모니터 — 회수·처분·안전성·공급중단·R&D 통합 평면화 (KD-IRIS v3)."""
    events = []
    try:
        # 회수 (NO 539)
        rr = fetch_recall(num_of_rows=30)
        for r in rr.get("items", [])[:15]:
            events.append({
                "type": "recall",
                "type_label": "회수·판매중지",
                "severity_level": "CRITICAL", "severity_color": "danger",
                "title": (r.get("PRDUCT") or "(제품 미상)")[:60],
                "summary": (r.get("RTRVL_RESN") or "")[:100],
                "entity": r.get("ENTRPS") or "",
                "date": (r.get("RECALL_COMMAND_DATE") or r.get("ENFRC_YN") or "")[:10],
            })
    except Exception as e:
        logger.warning(f"timeline recall 수집 실패: {e}")
    try:
        # 행정처분 (NO 564)
        dr = fetch_disciplinary(num_of_rows=20)
        for d in dr.get("items", [])[:10]:
            events.append({
                "type": "disciplinary",
                "type_label": "행정처분",
                "severity_level": "HIGH", "severity_color": "warn",
                "title": (d.get("ENTP_NAME") or "(업체 미상)")[:60],
                "summary": (d.get("VIOLATION_DETAIL") or d.get("VIOLATION_LAW") or "")[:100],
                "entity": d.get("ENTP_NAME") or "",
                "date": (d.get("DSPS_DCSNDT") or "")[:10],
            })
    except Exception as e:
        logger.warning(f"timeline disc 수집 실패: {e}")
    try:
        # 안전성서한 (NO 547)
        sl = fetch_drug_safety_letter(num_of_rows=15)
        for s in sl.get("items", [])[:8]:
            events.append({
                "type": "safety_letter",
                "type_label": "안전성서한",
                "severity_level": "HIGH", "severity_color": "warn",
                "title": (s.get("TITLE") or "")[:60],
                "summary": (s.get("SUMRY_CONT") or "")[:100],
                "entity": s.get("PBANC_DIVS_NM") or "",
                "date": (s.get("PBANC_YMD") or "")[:10],
            })
    except Exception as e:
        logger.warning(f"timeline safety 수집 실패: {e}")
    try:
        # 공급중단 (NO 534)
        ss = fetch_drug_supply_stop(num_of_rows=10)
        for s in ss.get("items", [])[:5]:
            events.append({
                "type": "supply_stop",
                "type_label": "공급중단",
                "severity_level": "CRITICAL", "severity_color": "danger",
                "title": (s.get("ITEM_NAME") or "(품목 미상)")[:60],
                "summary": (s.get("REPORT_PGS_CODE") or "") + " · " + (s.get("SUSPEND_REPORT_FLAG") or ""),
                "entity": s.get("ENTP_NAME") or "",
                "date": "",  # 응답에 날짜 필드명 불명확
            })
    except Exception as e:
        logger.warning(f"timeline supply 수집 실패: {e}")
    try:
        # 검사부적합 (NO 535)
        from ..api_extras import fetch_food_inspect
        fi = fetch_food_inspect(num_of_rows=15)
        for item in fi.get("items", [])[:8]:
            events.append({
                "type": "inspection_fail",
                "type_label": "검사부적합",
                "severity_level": "HIGH", "severity_color": "warn",
                "title": (item.get("PRDUCT") or "(품목 미상)")[:60],
                "summary": (item.get("IMPROPT_ITM") or "")[:100],
                "entity": item.get("ENTRPS") or "",
                "date": (item.get("REGIST_DT") or "")[:10],
            })
    except Exception as e:
        logger.warning(f"timeline inspection 수집 실패: {e}")
    try:
        # 신규 허가 (NO 140) — 최신 등재 의약품 5건
        ap = fetch_approval(num_of_rows=5)
        for item in ap.get("items", [])[:5]:
            events.append({
                "type": "permit_new",
                "type_label": "신규 허가",
                "severity_level": "LOW", "severity_color": "info",
                "title": (item.get("ITEM_NAME") or "(품목 미상)")[:60],
                "summary": (item.get("SPCLTY_PBLC") or "") + " · " + (item.get("PRDUCT_TYPE") or ""),
                "entity": item.get("ENTP_NAME") or "",
                "date": (item.get("ITEM_PERMIT_DATE") or "")[:8],
            })
    except Exception as e:
        logger.warning(f"timeline permit_new 수집 실패: {e}")

    # ────────────────────────────────────────────────────────────────────
    # R&D 이벤트 4종 — Phase 4에서 추가 (NO 566 임상 · NO 561 특허 · NO 554 재심사 · NO 556 재평가)
    # ────────────────────────────────────────────────────────────────────
    try:
        from ..api_extras import fetch_drug_clinical
        items = (fetch_drug_clinical(num_of_rows=10).get("items") or [])[:6]
        for it in items:
            events.append({
                "type": "clinical_new",
                "type_label": "임상시험 신규",
                "severity_level": "LOW", "severity_color": "info",
                "title": (it.get("GOODS_NAME") or "(임상 미상)")[:60],
                "summary": (it.get("CLINIC_EXAM_TITLE") or "")[:100],
                "entity": it.get("APPLY_ENTP_NAME") or "",
                "date": "",
            })
    except Exception as e:
        logger.warning(f"timeline clinical_new 수집 실패: {e}")

    try:
        from ..api_extras import fetch_drug_patent
        items = (fetch_drug_patent(num_of_rows=10).get("items") or [])[:6]
        for it in items:
            events.append({
                "type": "patent_new",
                "type_label": "특허 등록",
                "severity_level": "LOW", "severity_color": "info",
                "title": (it.get("ITEM_NAME") or it.get("PAT_NO") or "(특허 미상)")[:60],
                "summary": (it.get("INGR_NAME") or "")[:60] + (" · 만료 " + it.get("EXPIRE_DATE") if it.get("EXPIRE_DATE") else ""),
                "entity": it.get("ENTP_NAME") or "",
                "date": (it.get("EXPIRE_DATE") or "")[:10],
            })
    except Exception as e:
        logger.warning(f"timeline patent_new 수집 실패: {e}")

    try:
        from ..api_extras import fetch_drug_review
        items = (fetch_drug_review(num_of_rows=10).get("items") or [])[:5]
        for it in items:
            events.append({
                "type": "review_due",
                "type_label": "재심사 예정",
                "severity_level": "LOW", "severity_color": "info",
                "title": (it.get("ITEM_NAME") or "(품목 미상)")[:60],
                "summary": "재심사 일정",
                "entity": it.get("ENTP_NAME") or "",
                "date": (it.get("REJDGE_DT") or "")[:10],
            })
    except Exception as e:
        logger.warning(f"timeline review_due 수집 실패: {e}")

    try:
        from ..api_extras import fetch_drug_reeval
        items = (fetch_drug_reeval(num_of_rows=10).get("items") or [])[:5]
        for it in items:
            events.append({
                "type": "reeval_done",
                "type_label": "재평가 완료",
                "severity_level": "LOW", "severity_color": "info",
                "title": (it.get("ITEM_NAME") or "(품목 미상)")[:60],
                "summary": "재평가 결과 공시",
                "entity": it.get("ENTP_NAME") or "",
                "date": (it.get("REVAL_DT") or "")[:10],
            })
    except Exception as e:
        logger.warning(f"timeline reeval_done 수집 실패: {e}")

    # 심각도 + 시간순 정렬 (CRITICAL → HIGH → LOW, 같은 등급 안에서는 날짜 내림차순)
    from .severity import SEVERITY_ORDER
    from datetime import date as _date
    today = _date.today().isoformat()
    for e in events:
        e["is_new"] = bool(e.get("date")) and e["date"] == today
        if "type_label" not in e:
            e["type_label"] = e.get("type", "기타")
        if "in_watchlist" not in e:
            e["in_watchlist"] = False
    events.sort(key=lambda e: (
        SEVERITY_ORDER.get(e.get("severity_level", "NONE"), 3),
        -1 * int((e.get("date") or "0").replace("-", "") or 0),
    ))

    # severity 카운트 (KD-IRIS v3: CRITICAL/HIGH/LOW 3단계)
    severity_counts = {
        "critical": sum(1 for e in events if e.get("severity_level") == "CRITICAL"),
        "high":     sum(1 for e in events if e.get("severity_level") == "HIGH"),
        "low":      sum(1 for e in events if e.get("severity_level") == "LOW"),
    }
    type_counts = {}
    for e in events:
        type_counts[e["type"]] = type_counts.get(e["type"], 0) + 1

    return render_template(
        "app/monitor.html",
        active_page="monitor",
        events=events[:30],
        total_count=len(events),
        severity_counts=severity_counts,
        type_counts=type_counts,
    )


@app_bp.route("/copilot")
def copilot():
    """AI 코파일럿 — MISO drug_consult_agent 프록시 + 자사 현황 맥락 주입 (v2 §6)."""
    own = {"total": 0, "matched": 0, "critical": 0, "unconfirmed": 0, "review_due_90": 0}
    try:
        from .qa import lookup as qa_lookup
        own = qa_lookup.own_summary()
    except Exception as e:
        logger.debug(f"코파일럿 자사 요약 실패: {e}")
    return render_template("app/copilot.html", active_page="copilot", own=own)


@app_bp.route("/home")
def home():
    """오늘의 브리핑 — Phase R-2 (R2-6) 에서 풀 구현 예정. 현재는 placeholder."""
    return render_template("app/home.html", active_page="home")


# ────────────────────────────────────────────────────────────────────────────
# 워치리스트 — 사용자 등록 항목 (JSON 영속화)
# ────────────────────────────────────────────────────────────────────────────

@app_bp.route("/watchlist")
def watchlist():
    """워치리스트 페이지 — 등록 항목 목록 + 식약처 이벤트 매칭."""
    from . import watchlist_store
    from . import watchlist_match
    entries = watchlist_store.list_entries()
    # 항목별 식약처 이벤트 매칭 (캐시 5분 TTL — 같은 페이지 새로고침 시 API 재호출 안 함)
    matches = watchlist_match.match_for_entries(entries)
    alert_total = watchlist_match.total_alert_count(matches)
    return render_template(
        "app/watchlist.html",
        active_page="watchlist",
        entries=entries,
        matches=matches,
        alert_total=alert_total,
        flash_msg=request.args.get("msg") or "",
    )


@app_bp.route("/watchlist/add", methods=["POST"])
def watchlist_add():
    """워치리스트 항목 추가."""
    from . import watchlist_store
    query = (request.form.get("query") or "").strip()
    label = (request.form.get("label") or "").strip() or query
    kind = (request.form.get("kind") or "product").strip()
    note = (request.form.get("note") or "").strip()
    redirect_to = request.form.get("next") or "/app/watchlist"
    if not query:
        return redirect(redirect_to + ("?msg=쿼리를_입력하세요" if "?" not in redirect_to else "&msg=쿼리를_입력하세요"))
    watchlist_store.add_entry(query=query, label=label, kind=kind, note=note)
    sep = "&" if "?" in redirect_to else "?"
    return redirect(f"{redirect_to}{sep}msg=추가됨")


@app_bp.route("/watchlist/upload", methods=["POST"])
def watchlist_upload():
    """워치리스트 Excel 일괄 업로드 (v3 R3-5) — 품목명/ITEM_NAME 컬럼 1열 읽기."""
    from . import watchlist_store
    f = request.files.get("file")
    if not f or not f.filename:
        return redirect("/app/watchlist?msg=파일을_선택하세요")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return redirect("/app/watchlist?msg=빈_파일")
        # 헤더에서 품목명 컬럼 찾기 (ITEM_NAME / 품목명 / 제품명), 없으면 1열
        header = [str(c or "").strip() for c in rows[0]]
        name_idx = 0
        for i, h in enumerate(header):
            if h in ("ITEM_NAME", "품목명", "제품명", "품목"):
                name_idx = i
                break
        has_header = any(h in ("ITEM_NAME", "품목명", "제품명", "품목") for h in header)
        data_rows = rows[1:] if has_header else rows
        added = 0
        for r in data_rows:
            if name_idx >= len(r):
                continue
            name = str(r[name_idx] or "").strip()
            if name:
                watchlist_store.add_entry(query=name, label=name, kind="product",
                                          note="Excel 일괄 업로드")
                added += 1
        return redirect(f"/app/watchlist?msg={added}건_추가됨")
    except Exception as e:
        logger.warning(f"워치리스트 업로드 실패: {e}")
        return redirect("/app/watchlist?msg=업로드_실패")


@app_bp.route("/watchlist/<entry_id>/delete", methods=["POST"])
def watchlist_delete(entry_id):
    """워치리스트 항목 삭제."""
    from . import watchlist_store
    watchlist_store.delete_entry(entry_id)
    return redirect("/app/watchlist?msg=삭제됨")


# ────────────────────────────────────────────────────────────────────────────
# 부서별 워크스페이스 — 7개 부서 동적 페이지
# ────────────────────────────────────────────────────────────────────────────

# 이벤트 종류별 fetch + 정규화 함수. watchlist_match와 동일한 캐시 활용.
def _fetch_workspace_events(kinds):
    """
    워크스페이스가 관심 갖는 이벤트 종류 list 에 대해
    글로벌 API 결과를 fetch(캐시 적용)하고 정규화된 이벤트 리스트로 반환.

    kinds: ["recall", "safety", "sanction", "supply", "new_permit",
            "inspection", "food_recall", "gmp", "clinical"]
    """
    from . import watchlist_match  # 캐시 _cached 재사용
    from .severity import get_severity

    out = {}  # kind -> list of normalized events

    if "recall" in kinds:
        items = watchlist_match._cached("recall", fetch_recall, num_of_rows=50)
        out["recall"] = [watchlist_match._normalize_event("recall", r) for r in items[:8]]

    if "safety" in kinds:
        from ..api_extras import fetch_drug_safety_letter
        items = watchlist_match._cached("safety", fetch_drug_safety_letter, num_of_rows=50)
        out["safety"] = [watchlist_match._normalize_event("safety", s) for s in items[:8]]

    if "sanction" in kinds:
        items = watchlist_match._cached("sanction", fetch_disciplinary, num_of_rows=50)
        out["sanction"] = [watchlist_match._normalize_event("sanction", d) for d in items[:8]]

    if "supply" in kinds:
        from ..api_extras import fetch_drug_supply_stop
        items = watchlist_match._cached("supply", fetch_drug_supply_stop, num_of_rows=50)
        out["supply"] = [watchlist_match._normalize_event("supply", u) for u in items[:8]]

    if "new_permit" in kinds:
        try:
            items = (fetch_approval(num_of_rows=20).get("items") or [])[:8]
            out["new_permit"] = [{
                "type": "new_permit", "type_label": "신규 허가", "severity_level": "LOW", "severity_color": "info",
                "title": (it.get("ITEM_NAME") or "")[:60],
                "summary": (it.get("ENTP_NAME") or "") + " · " + (it.get("SPCLTY_PBLC") or ""),
                "date": (it.get("ITEM_PERMIT_DATE") or "")[:8],
            } for it in items]
        except Exception:
            out["new_permit"] = []

    if "inspection" in kinds:
        try:
            from ..api_extras import fetch_food_inspect
            items = (fetch_food_inspect(num_of_rows=20).get("items") or [])[:8]
            out["inspection"] = [{
                "type": "inspection", "type_label": "검사부적합", "severity_level": "HIGH", "severity_color": "warn",
                "title": (it.get("PRDUCT") or "")[:60],
                "summary": (it.get("IMPROPT_ITM") or "")[:80] + " · " + (it.get("ENTRPS") or ""),
                "date": (it.get("REGIST_DT") or "")[:10],
            } for it in items]
        except Exception:
            out["inspection"] = []

    if "food_recall" in kinds:
        try:
            from ..api_extras import fetch_food_recall_import
            items = (fetch_food_recall_import(num_of_rows=20).get("items") or [])[:8]
            out["food_recall"] = [{
                "type": "food_recall", "type_label": "수입식품 회수", "severity_level": "HIGH", "severity_color": "warn",
                "title": (it.get("PRDT_NM") or "")[:60],
                "summary": (it.get("RECL_RSN") or "") + " · " + (it.get("CLNT_BSSH_NM") or ""),
                "date": (it.get("CIRC_PRD") or "")[:10],
            } for it in items]
        except Exception:
            out["food_recall"] = []

    if "gmp" in kinds:
        try:
            from ..api_extras import fetch_drug_gmp
            items = (fetch_drug_gmp(num_of_rows=20).get("items") or [])[:6]
            out["gmp"] = [{
                "type": "gmp", "type_label": "GMP 적합판정", "severity_level": "LOW", "severity_color": "info",
                "title": (it.get("ITEM_NAME") or it.get("ENTP_NAME") or "")[:60],
                "summary": "GMP 적합판정 · " + (it.get("ENTP_NAME") or ""),
                "date": (it.get("IUS_DT") or "")[:10],
            } for it in items]
        except Exception:
            out["gmp"] = []

    if "clinical" in kinds:
        try:
            from ..api_extras import fetch_drug_clinical
            items = watchlist_match._cached("clinical", fetch_drug_clinical, num_of_rows=50)[:6]
            out["clinical"] = [{
                "type": "clinical", "type_label": "임상시험", "severity_level": "LOW", "severity_color": "info",
                "title": (it.get("GOODS_NAME") or "")[:60],
                "summary": (it.get("CLINIC_EXAM_TITLE") or "")[:80],
                "date": "",
            } for it in items]
        except Exception:
            out["clinical"] = []

    if "review" in kinds:
        try:
            from ..api_extras import fetch_drug_review
            items = watchlist_match._cached("review", fetch_drug_review, num_of_rows=50)[:6]
            out["review"] = [{
                "type": "review", "type_label": "재심사", "severity_level": "HIGH", "severity_color": "warn",
                "title": (it.get("ITEM_NAME") or "")[:60],
                "summary": (it.get("ENTP_NAME") or ""),
                "date": (it.get("REJDGE_DT") or it.get("REVAL_DT") or "")[:10],
            } for it in items]
        except Exception:
            out["review"] = []

    if "reeval" in kinds:
        try:
            from ..api_extras import fetch_drug_reeval
            items = watchlist_match._cached("reeval", fetch_drug_reeval, num_of_rows=50)[:6]
            out["reeval"] = [{
                "type": "reeval", "type_label": "재평가", "severity_level": "HIGH", "severity_color": "warn",
                "title": (it.get("ITEM_NAME") or "")[:60],
                "summary": (it.get("ENTP_NAME") or ""),
                "date": (it.get("REVAL_DT") or "")[:10],
            } for it in items]
        except Exception:
            out["reeval"] = []

    if "patent" in kinds:
        try:
            from ..api_extras import fetch_drug_patent
            items = watchlist_match._cached("patent", fetch_drug_patent, num_of_rows=50)[:6]
            out["patent"] = [{
                "type": "patent", "type_label": "특허 등록", "severity_level": "LOW", "severity_color": "info",
                "title": (it.get("ITEM_NAME") or it.get("PAT_NO") or "")[:60],
                "summary": (it.get("INGR_NAME") or "") + " · " + (it.get("ENTP_NAME") or ""),
                "date": (it.get("EXPIRE_DATE") or "")[:10],
            } for it in items]
        except Exception:
            out["patent"] = []

    if "bioeq" in kinds:
        try:
            from ..api_extras import fetch_drug_bioeq
            items = watchlist_match._cached("bioeq", fetch_drug_bioeq, num_of_rows=50)[:6]
            out["bioeq"] = [{
                "type": "bioeq", "type_label": "생동성인정", "severity_level": "LOW", "severity_color": "info",
                "title": (it.get("ITEM_NAME") or "")[:60],
                "summary": (it.get("INGR_KOR_NAME") or "") + " · " + (it.get("ENTP_NAME") or ""),
                "date": "",
            } for it in items]
        except Exception:
            out["bioeq"] = []

    if "supply_lack" in kinds:
        try:
            from ..api_extras import fetch_drug_supply_lack
            items = watchlist_match._cached("supply_lack", fetch_drug_supply_lack, num_of_rows=50)[:6]
            out["supply_lack"] = [{
                "type": "supply_lack", "type_label": "공급부족", "severity_level": "HIGH", "severity_color": "warn",
                "title": (it.get("ITEM_NAME") or "")[:60],
                "summary": (it.get("ENTP_NAME") or ""),
                "date": "",
            } for it in items]
        except Exception:
            out["supply_lack"] = []

    if "haccp_smart" in kinds:
        try:
            from ..api_extras import fetch_haccp_smart
            items = watchlist_match._cached("haccp_smart", fetch_haccp_smart, num_of_rows=50)[:6]
            out["haccp_smart"] = [{
                "type": "haccp_smart", "type_label": "스마트HACCP", "severity_level": "LOW", "severity_color": "info",
                "title": (it.get("BSSH_NM") or "")[:60],
                "summary": (it.get("ADDR") or "") + " · " + (it.get("PRMS_NO") or ""),
                "date": (it.get("PRMS_DT") or "")[:10],
            } for it in items]
        except Exception:
            out["haccp_smart"] = []

    return out


@app_bp.route("/workspace")
def workspace_index():
    """워크스페이스 단일 진입 — 부서 미지정 시 RA 로 redirect."""
    return redirect(url_for("app.workspace", dept_id="ra"), code=302)


@app_bp.route("/workspace/<dept_id>")
def workspace(dept_id):
    """부서별 워크스페이스 — workspaces_config 메타데이터 + 실 이벤트."""
    from .workspaces_config import by_id
    from . import watchlist_store, watchlist_match

    ws = by_id(dept_id)
    if not ws:
        from flask import abort
        abort(404)

    # 부서 관심 이벤트 fetch + 정규화 (캐시 사용)
    events_by_kind = _fetch_workspace_events(ws["event_kinds"])

    # KPI 값 계산
    kpi_values = {}
    for spec in ws["kpi_specs"]:
        kind = spec["kind"]
        key  = spec["key"]
        if kind == "count":
            if key in events_by_kind:
                kpi_values[key] = len(events_by_kind[key])
            elif key == "all_alerts":
                # 경영진 워크스페이스 — 모든 관심 이벤트의 총합
                kpi_values[key] = sum(len(v) for v in events_by_kind.values())
            else:
                kpi_values[key] = 0
        elif kind == "watchlist_count":
            try:
                kpi_values[key] = watchlist_store.count()
            except Exception:
                kpi_values[key] = 0
        elif kind == "static":
            kpi_values[key] = spec.get("value", "—")
        else:
            kpi_values[key] = "—"

    # 전체 이벤트 평면화 + 날짜 내림차순
    flat_events = []
    for evs in events_by_kind.values():
        flat_events.extend(evs)
    flat_events.sort(key=lambda e: e.get("date", ""), reverse=True)

    # RA 워크스페이스 — 자사 재심사 D-Day 캘린더 (v3 R3-1, master_item.reexam_date 활용)
    reexam_calendar = None
    if dept_id == "ra":
        reexam_calendar = _reexam_dday()

    return render_template(
        "app/workspace.html",
        active_workspace=dept_id,
        workspace=ws,
        kpi_values=kpi_values,
        events=flat_events[:15],
        events_by_kind=events_by_kind,
        reexam_calendar=reexam_calendar,
    )


def _reexam_dday():
    """자사 마스터 재심사 예정일 → D-Day 버킷 (D-30/D-90/D-365)."""
    import datetime
    try:
        from .qa import db as qadb
        today = datetime.date.today()
        # reexam_date 는 "시작~종료" 범위 형식 다수 → 종료일(마감) 기준. 소량이라 전체 fetch 후 파싱.
        rows = qadb.query(
            """SELECT item_code, item_name, reexam_date FROM master_item
               WHERE active=1 AND reexam_date IS NOT NULL AND reexam_date != ''""")
        buckets = {"d30": [], "d90": [], "d365": []}
        for r in rows:
            raw = r["reexam_date"] or ""
            end = raw.split("~")[-1]                 # 범위면 종료일, 단일이면 그대로
            rd = "".join(ch for ch in end if ch.isdigit())[:8]
            if len(rd) != 8:
                continue
            try:
                exp = datetime.date(int(rd[:4]), int(rd[4:6]), int(rd[6:8]))
            except ValueError:
                continue
            dday = (exp - today).days
            if dday < 0:
                continue  # 이미 만료(재심사 완료)
            item = {"item_code": r["item_code"], "item_name": r["item_name"],
                    "reexam_date": end.strip(), "dday": dday}
            if dday <= 30:
                buckets["d30"].append(item)
            elif dday <= 90:
                buckets["d90"].append(item)
            elif dday <= 365:
                buckets["d365"].append(item)
        for b in buckets.values():
            b.sort(key=lambda x: x["dday"])
        return {"d30": buckets["d30"], "d90": buckets["d90"], "d365": buckets["d365"][:10],
                "total": sum(len(b) for b in buckets.values())}
    except Exception as e:
        logger.debug(f"재심사 D-Day 계산 실패: {e}")
        return None


# ────────────────────────────────────────────────────────────────────────────
# 리포트 — 사전 정의 리포트 카탈로그
# ────────────────────────────────────────────────────────────────────────────

# 카드 그리드에 표시할 사전 정의 리포트 5개.
# 실제 PDF/Excel 다운로드는 Phase 2.5 — 현재는 카드 클릭 시 안내 페이지로 폴백.
_REPORT_CATALOG = [
    {
        "id": "monthly-recall",
        "title": "월간 회수·판매중지 동향",
        "summary": "지난 30일간 식약처 NO 539 회수 데이터를 도메인·사유별로 집계. 자사 영향 가능성 라벨 포함.",
        "type": "회수",
        "kind_class": "danger",
        "cadence": "월간",
        "last_updated": "2026-05-25",
        "pages": 12,
        "data_sources": ["NO 539", "NO 547"],
    },
    {
        "id": "quarterly-rnd",
        "title": "분기 R&D 인사이트",
        "summary": "특허·임상시험·대조약·생동성 7개 API 매칭 결과를 성분군별로 정리. 경쟁사 파이프라인 변동 포함.",
        "type": "R&D",
        "kind_class": "brand",
        "cadence": "분기",
        "last_updated": "2026-04-30",
        "pages": 28,
        "data_sources": ["NO 561", "NO 566", "NO 484", "NO 485", "NO 552", "NO 562", "NO 557"],
    },
    {
        "id": "company-safety",
        "title": "자사 안전성 이벤트 요약",
        "summary": "광동제약 자사 품목에 대한 안전성서한·DUR·회수·공급중단 통합 타임라인. RA·QA 공유용.",
        "type": "안전성",
        "kind_class": "warn",
        "cadence": "주간",
        "last_updated": "2026-05-26",
        "pages": 8,
        "data_sources": ["NO 547", "NO 531", "NO 539", "NO 534"],
    },
    {
        "id": "gmp-status",
        "title": "GMP · 심사 상태 리포트",
        "summary": "자사 제조소 GMP 적합판정·재심사·재평가·DMF 일정과 만료 위험 품목. 매월 1일 갱신.",
        "type": "품질",
        "kind_class": "ok",
        "cadence": "월간",
        "last_updated": "2026-05-01",
        "pages": 6,
        "data_sources": ["NO 132", "NO 142", "NO 483", "NO 554", "NO 556"],
    },
    {
        "id": "competitor-sanction",
        "title": "경쟁사 행정처분 동향",
        "summary": "한미·동아·유한·제일 등 주요 경쟁사 행정처분(NO 564) 90일 추적. 영업 인사이트 카드 포함.",
        "type": "행정처분",
        "kind_class": "warn",
        "cadence": "분기",
        "last_updated": "2026-05-20",
        "pages": 14,
        "data_sources": ["NO 564", "NO 539"],
    },
]


@app_bp.route("/reports")
def reports():
    """리포트 카탈로그 페이지."""
    return render_template(
        "app/reports.html",
        active_page="reports",
        reports=_REPORT_CATALOG,
    )


# ────────────────────────────────────────────────────────────────────────────
# 검사부적합 트래커 — NO 535 전용 페이지 (Phase 4)
# ────────────────────────────────────────────────────────────────────────────

@app_bp.route("/inspections")
def inspections():
    """식품 검사부적합 트래커 — NO 535 전용 필터/조회 페이지."""
    from ..api_extras import fetch_food_inspect
    from datetime import datetime, timedelta

    q = (request.args.get("q") or "").strip()
    # 글로벌 fetch (캐시 활용)
    try:
        from . import watchlist_match
        all_items = watchlist_match._cached("food_inspect_track", fetch_food_inspect, num_of_rows=200)
    except Exception:
        all_items = []

    # 검색어 필터 (없으면 전체)
    fields = ["PRDUCT", "ENTRPS", "IMPROPT_ITM", "RAWMTRL_NM", "BSSH_NM"]
    items = _filter_items(all_items, q, fields) if q else all_items

    # KPI 계산
    # 전체 부적합 / 최근 30일 / 자사(광동) 영향 / 부적합 사유 TOP
    today_str = datetime.now().strftime("%Y%m%d")
    thirty_days_ago = (datetime.now() - timedelta(days=30)).strftime("%Y%m%d")

    recent_30 = 0
    our_company_hits = 0
    class_counter = {}        # 사유 분류 (잔류농약/중금속/미생물/이물·기타)
    maker_counter = {}        # 제조소별 부적합 카운트
    for it in items:
        date = (it.get("REGIST_DT") or it.get("DSPS_DCSNDT") or "")[:8]
        if date and date >= thirty_days_ago.replace("-", ""):
            recent_30 += 1
        entrps = (it.get("ENTRPS") or "") + " " + (it.get("BSSH_NM") or "")
        if "광동" in entrps:
            our_company_hits += 1
        # 부적합 사유 분류
        cls = _classify_reason(it.get("IMPROPT_ITM") or "")
        class_counter[cls] = class_counter.get(cls, 0) + 1
        # 제조소 순위
        maker = (it.get("ENTRPS") or it.get("BSSH_NM") or "").strip()
        if maker:
            maker_counter[maker] = maker_counter.get(maker, 0) + 1

    total = len(items)
    # 분류 차트 (비율 포함, 큰 순)
    class_chart = [
        {"label": k, "count": v, "pct": round(v / total * 100) if total else 0}
        for k, v in sorted(class_counter.items(), key=lambda x: x[1], reverse=True)
    ]
    # 제조소 TOP5 (5건↑ 위험 표시)
    maker_top = [
        {"name": k, "count": v, "risk": v >= 5}
        for k, v in sorted(maker_counter.items(), key=lambda x: x[1], reverse=True)[:5]
    ]

    kpi = {
        "total": total,
        "recent_30": recent_30,
        "our_company": our_company_hits,
        "class_chart": class_chart,
        "maker_top": maker_top,
    }

    # 광동 매칭 항목 상단 고정
    items_sorted = sorted(items, key=lambda it: 0 if "광동" in ((it.get("ENTRPS") or "") + (it.get("BSSH_NM") or "")) else 1)

    return render_template(
        "app/inspections.html",
        active_page="inspections",
        query=q,
        items=items_sorted[:50],
        kpi=kpi,
    )


# 검사부적합 사유 분류 (v3 R3-4)
_REASON_RULES = [
    ("잔류농약", ("농약", "잔류농약", "피프로닐", "글리포세이트", "클로르")),
    ("중금속", ("중금속", "납", "카드뮴", "비소", "수은", "크롬")),
    ("미생물", ("대장균", "세균", "미생물", "살모넬라", "리스테리아", "황색포도상", "효모", "곰팡이")),
    ("기준초과", ("기준", "초과", "함량", "산가", "과산화물", "타르색소", "보존료", "사카린", "벤조피렌")),
    ("이물·기타", ()),
]


def _classify_reason(text: str) -> str:
    """부적합 사유 텍스트 → 분류 (잔류농약/중금속/미생물/기준초과/이물·기타)."""
    t = text or ""
    for label, kws in _REASON_RULES:
        if kws and any(k in t for k in kws):
            return label
    return "이물·기타"


# ────────────────────────────────────────────────────────────────────────────
# Excel 내보내기 공통 라우트 (v3 R3-7)
# ────────────────────────────────────────────────────────────────────────────
def _rows_for_export(resource: str):
    """resource → (sheet_title, headers, rows). openpyxl 워크북 빌드용."""
    if resource == "qa_master":
        from .qa import db as qadb
        data = qadb.query(
            """SELECT item_code, item_name, category, sub_category, gubun, self_or_consign,
                      item_seq, permit_no, reexam_date, atc_code, edi_code,
                      enrich_status, enrich_confidence, source_api
               FROM master_item WHERE active=1 ORDER BY item_code""")
        headers = ["품목코드", "품목명", "전문분류", "품목분류", "품목구분", "자사/위탁",
                   "ITEM_SEQ", "허가번호", "재심사일", "ATC", "EDI", "enrich상태", "confidence", "출처"]
        rows = [[d.get(k) for k in
                 ("item_code", "item_name", "category", "sub_category", "gubun", "self_or_consign",
                  "item_seq", "permit_no", "reexam_date", "atc_code", "edi_code",
                  "enrich_status", "enrich_confidence", "source_api")] for d in data]
        return "자사품목마스터", headers, rows

    if resource == "qa_events":
        from .qa import db as qadb
        data = qadb.query(
            """SELECT event_date, event_type, severity, impact_level, title, entity, status, summary
               FROM event ORDER BY detected_at DESC""")
        headers = ["발생일", "유형", "심각도", "자사영향", "제목", "업체", "상태", "요약"]
        rows = [[d.get(k) for k in
                 ("event_date", "event_type", "severity", "impact_level", "title", "entity", "status", "summary")]
                for d in data]
        return "QA이벤트", headers, rows

    if resource == "inspections":
        from ..api_extras import fetch_food_inspect
        try:
            from . import watchlist_match
            items = watchlist_match._cached("food_inspect_track", fetch_food_inspect, num_of_rows=200)
        except Exception:
            items = []
        headers = ["제품명", "업체", "부적합사유", "분류", "원재료", "등록일"]
        rows = [[it.get("PRDUCT"), it.get("ENTRPS"), it.get("IMPROPT_ITM"),
                 _classify_reason(it.get("IMPROPT_ITM") or ""), it.get("RAWMTRL_NM"),
                 (it.get("REGIST_DT") or "")[:8]] for it in items]
        return "검사부적합", headers, rows

    return None


@app_bp.route("/export/<resource>")
def export_xlsx(resource):
    """리소스를 xlsx로 내보내기. resource ∈ {qa_master, qa_events, inspections}."""
    import datetime
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({"error": "openpyxl 미설치"}), 500

    bundle = _rows_for_export(resource)
    if not bundle:
        return jsonify({"error": f"알 수 없는 resource: {resource}"}), 404
    title, headers, rows = bundle

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for r in rows:
        ws.append([("" if v is None else str(v)) for v in r])
    # 컬럼 폭 자동(간이)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(10, min(40, len(str(h)) + 6))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"KD-IRIS_{resource}_{datetime.date.today().isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
